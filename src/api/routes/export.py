"""
API Routes für Daten-Export (Excel, CSV, JSON).
Ermöglicht Export von Suchergebnissen und gefilterten Entscheidungen.
"""

from typing import List, Any
from io import BytesIO, StringIO
from datetime import datetime
import csv
import json
import structlog

from fastapi import APIRouter, Depends, Query, HTTPException, status
from fastapi.responses import StreamingResponse, Response
from sqlalchemy import select, and_, or_
from sqlalchemy.ext.asyncio import AsyncSession

try:
    import pandas as pd

    PANDAS_AVAILABLE = True
except ImportError:
    PANDAS_AVAILABLE = False

from src.database import Decision
from src.api import schemas
from src.api.deps import get_db, get_decision_filter
from src.config import settings

logger = structlog.get_logger()

router = APIRouter()


# =============================================================================
# HILFSFUNKTIONEN
# =============================================================================


async def get_filtered_decisions(
    db: AsyncSession, filters: schemas.DecisionFilter, limit: int = None
) -> List[Decision]:
    """Holt gefilterte Entscheidungen aus der Datenbank."""

    query = select(Decision)
    conditions = []

    if filters.source:
        conditions.append(Decision.source == filters.source)

    if filters.court:
        conditions.append(Decision.court.ilike(f"%{filters.court}%"))

    if filters.gdpr_article:
        conditions.append(Decision.gdpr_articles.contains([filters.gdpr_article]))

    if filters.date_from:
        conditions.append(Decision.decision_date >= filters.date_from)

    if filters.date_to:
        conditions.append(Decision.decision_date <= filters.date_to)

    if filters.keyword:
        conditions.append(
            or_(
                Decision.title.ilike(f"%{filters.keyword}%"),
                Decision.keywords.contains([filters.keyword]),
            )
        )

    if conditions:
        query = query.where(and_(*conditions))

    # Sortierung
    query = query.order_by(Decision.decision_date.desc())

    # Limit anwenden
    if limit:
        query = query.limit(limit)

    result = await db.execute(query)
    return result.scalars().all()


def decisions_to_dataframe(decisions: List[Decision]) -> Any:
    """Konvertiert Entscheidungen zu Pandas DataFrame."""

    if not PANDAS_AVAILABLE:
        raise ImportError("Pandas ist nicht installiert")

    data = []
    for d in decisions:
        data.append(
            {
                "ID": str(d.id),
                "Aktenzeichen": d.case_number,
                "Titel": d.title,
                "Gericht": d.court,
                "Datum": d.decision_date.strftime("%Y-%m-%d") if d.decision_date else "",
                "DSGVO-Artikel": ", ".join(d.gdpr_articles) if d.gdpr_articles else "",
                "Keywords": ", ".join(d.keywords) if d.keywords else "",
                "Leitsatz": (d.leitsatz[:200] + "...") if d.leitsatz else "",
                "Quelle": d.source,
                "URL": d.source_url,
                "Anonymisiert": "Ja" if d.anonymization_applied else "Nein",
                "Qualität": d.avg_quality_score if d.avg_quality_score else "",
            }
        )

    return pd.DataFrame(data)


# =============================================================================
# EXCEL EXPORT
# =============================================================================


@router.get(
    "/excel",
    summary="Excel-Export",
    description="""
    Exportiert gefilterte Entscheidungen als Excel-Datei.
    
    **Features:**
    - Formatierte Tabelle mit Autofilter
    - Breite Spalten für bessere Lesbarkeit
    - Farbkodierung für verschiedene Gerichte
    - Maximal 10.000 Zeilen pro Export
    """,
    responses={
        200: {
            "description": "Excel-Datei",
            "content": {"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": {}},
        }
    },
)
async def export_excel(
    db: AsyncSession = Depends(get_db),
    filters: schemas.DecisionFilter = Depends(get_decision_filter),
    limit: int = Query(default=1000, le=settings.export_max_rows),
) -> StreamingResponse:
    """Exportiert Entscheidungen als Excel-Datei."""

    if not PANDAS_AVAILABLE:
        raise HTTPException(
            status_code=status.HTTP_501_NOT_IMPLEMENTED,
            detail="Excel-Export nicht verfügbar. Pandas/openpyxl nicht installiert.",
        )

    # Daten abrufen
    decisions = await get_filtered_decisions(db, filters, limit)

    if not decisions:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Keine Entscheidungen gefunden"
        )

    # DataFrame erstellen
    df = decisions_to_dataframe(decisions)

    # Excel generieren
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Hauptblatt
        df.to_excel(writer, sheet_name="Entscheidungen", index=False)

        # Formatierung
        worksheet = writer.sheets["Entscheidungen"]

        # Spaltenbreiten anpassen
        column_widths = {
            "A": 40,  # ID
            "B": 20,  # Aktenzeichen
            "C": 60,  # Titel
            "D": 20,  # Gericht
            "E": 12,  # Datum
            "F": 30,  # DSGVO-Artikel
            "G": 40,  # Keywords
            "H": 80,  # Leitsatz
            "I": 15,  # Quelle
            "J": 50,  # URL
            "K": 12,  # Anonymisiert
            "L": 10,  # Qualität
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Autofilter hinzufügen
        worksheet.auto_filter.ref = f"A1:L{len(df) + 1}"

        # Header formatieren
        from openpyxl.styles import Font, PatternFill

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")

        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.font = Font(bold=True, color="FFFFFF")

        # Bedingte Formatierung für Gerichte
        from openpyxl.formatting.rule import CellIsRule

        # BGH in Grün
        worksheet.conditional_formatting.add(
            f"D2:D{len(df) + 1}",
            CellIsRule(
                operator="containsText", formula=["BGH"], fill=PatternFill(bgColor="C6EFCE")
            ),
        )

        # OLG in Blau
        worksheet.conditional_formatting.add(
            f"D2:D{len(df) + 1}",
            CellIsRule(
                operator="containsText", formula=["OLG"], fill=PatternFill(bgColor="BDD7EE")
            ),
        )

    # Metadaten-Blatt hinzufügen
    with pd.ExcelWriter(output, engine="openpyxl", mode="a") as writer:
        metadata = pd.DataFrame(
            {
                "Information": ["Export-Datum", "Anzahl Entscheidungen", "Filter"],
                "Wert": [
                    datetime.now().strftime("%Y-%m-%d %H:%M"),
                    len(decisions),
                    str(filters.model_dump(exclude_none=True)),
                ],
            }
        )
        metadata.to_excel(writer, sheet_name="Metadaten", index=False)

    output.seek(0)

    logger.info(
        "excel_exported", count=len(decisions), filters=filters.model_dump(exclude_none=True)
    )

    # Dateiname mit Datum
    filename = f"dsr_decisions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =============================================================================
# CSV EXPORT
# =============================================================================


@router.get(
    "/csv",
    summary="CSV-Export",
    description="""
    Exportiert gefilterte Entscheidungen als CSV-Datei.
    
    **Features:**
    - UTF-8 Encoding mit BOM für Excel-Kompatibilität
    - Semikolon als Trennzeichen (deutsche Excel-Konvention)
    - Maximal 10.000 Zeilen pro Export
    """,
    responses={200: {"description": "CSV-Datei", "content": {"text/csv": {}}}},
)
async def export_csv(
    db: AsyncSession = Depends(get_db),
    filters: schemas.DecisionFilter = Depends(get_decision_filter),
    limit: int = Query(default=1000, le=settings.export_max_rows),
    delimiter: str = Query(default=";", description="CSV-Trennzeichen"),
) -> StreamingResponse:
    """Exportiert Entscheidungen als CSV-Datei."""

    # Daten abrufen
    decisions = await get_filtered_decisions(db, filters, limit)

    if not decisions:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Keine Entscheidungen gefunden"
        )

    # CSV generieren
    output = StringIO()

    # UTF-8 BOM für Excel
    output.write("\ufeff")

    # CSV Writer
    fieldnames = [
        "ID",
        "Aktenzeichen",
        "Titel",
        "Gericht",
        "Datum",
        "DSGVO-Artikel",
        "Keywords",
        "Leitsatz",
        "Quelle",
        "URL",
        "Anonymisiert",
    ]

    writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=delimiter)
    writer.writeheader()

    for d in decisions:
        writer.writerow(
            {
                "ID": str(d.id),
                "Aktenzeichen": d.case_number or "",
                "Titel": d.title,
                "Gericht": d.court or "",
                "Datum": d.decision_date.strftime("%Y-%m-%d") if d.decision_date else "",
                "DSGVO-Artikel": ", ".join(d.gdpr_articles) if d.gdpr_articles else "",
                "Keywords": ", ".join(d.keywords) if d.keywords else "",
                "Leitsatz": (d.leitsatz[:200] + "...") if d.leitsatz else "",
                "Quelle": d.source,
                "URL": d.source_url or "",
                "Anonymisiert": "Ja" if d.anonymization_applied else "Nein",
            }
        )

    output.seek(0)

    logger.info(
        "csv_exported",
        count=len(decisions),
        delimiter=delimiter,
        filters=filters.model_dump(exclude_none=True),
    )

    # Dateiname mit Datum
    filename = f"dsr_decisions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"

    return StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =============================================================================
# JSON EXPORT
# =============================================================================


@router.get(
    "/json",
    summary="JSON-Export",
    description="""
    Exportiert gefilterte Entscheidungen als JSON-Datei.
    
    **Features:**
    - Vollständige Daten inkl. strukturierter Felder
    - Pretty-Print für bessere Lesbarkeit
    - Maximal 10.000 Entscheidungen pro Export
    """,
    response_model=List[schemas.DecisionResponse],
)
async def export_json(
    db: AsyncSession = Depends(get_db),
    filters: schemas.DecisionFilter = Depends(get_decision_filter),
    limit: int = Query(default=100, le=settings.export_max_rows),
    pretty: bool = Query(default=True, description="Pretty-Print JSON"),
) -> Response:
    """Exportiert Entscheidungen als JSON-Datei."""

    # Daten abrufen
    decisions = await get_filtered_decisions(db, filters, limit)

    if not decisions:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="Keine Entscheidungen gefunden"
        )

    # JSON generieren
    data = {
        "metadata": {
            "export_date": datetime.now().isoformat(),
            "count": len(decisions),
            "filters": filters.model_dump(exclude_none=True),
            "version": "1.0",
        },
        "decisions": [
            {
                "id": str(d.id),
                "source": d.source,
                "source_id": d.source_id,
                "title": d.title,
                "case_number": d.case_number,
                "court": d.court,
                "decision_date": d.decision_date.isoformat() if d.decision_date else None,
                "gdpr_articles": d.gdpr_articles,
                "keywords": d.keywords,
                "leitsatz": d.leitsatz,
                "tenor": d.tenor,
                "full_text": d.full_text_anonymized if d.anonymization_applied else None,
                "source_url": d.source_url,
                "anonymization_applied": d.anonymization_applied,
                "quality_score": d.avg_quality_score,
                "rechtskraft_status": d.rechtskraft_status,
            }
            for d in decisions
        ],
    }

    # JSON String erstellen
    if pretty:
        json_str = json.dumps(data, ensure_ascii=False, indent=2)
    else:
        json_str = json.dumps(data, ensure_ascii=False)

    logger.info(
        "json_exported",
        count=len(decisions),
        pretty=pretty,
        filters=filters.model_dump(exclude_none=True),
    )

    # Dateiname mit Datum
    filename = f"dsr_decisions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"

    return Response(
        content=json_str,
        media_type="application/json",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


# =============================================================================
# STATISTIK-EXPORT
# =============================================================================


@router.get(
    "/stats",
    summary="Statistik-Export",
    description="Exportiert aggregierte Statistiken als JSON oder CSV",
)
async def export_statistics(
    format: str = Query(default="json", pattern="^(json|csv)$"), db: AsyncSession = Depends(get_db)
) -> Response:
    """Exportiert Statistiken."""

    # Statistiken sammeln
    from sqlalchemy import func

    # Gesamtanzahl
    total = await db.scalar(select(func.count(Decision.id)))

    # Nach Quelle
    by_source = await db.execute(
        select(Decision.source, func.count(Decision.id)).group_by(Decision.source)
    )

    # Nach Gericht (Top 10)
    by_court = await db.execute(
        select(Decision.court, func.count(Decision.id))
        .where(Decision.court.isnot(None))
        .group_by(Decision.court)
        .order_by(func.count(Decision.id).desc())
        .limit(10)
    )

    # DSGVO-Artikel (Top 10)
    dsr_stats = await db.execute(
        select(func.unnest(Decision.gdpr_articles).label("article"), func.count().label("count"))
        .group_by("article")
        .order_by("count DESC")
        .limit(10)
    )

    if format == "json":
        data = {
            "export_date": datetime.now().isoformat(),
            "total_decisions": total,
            "by_source": {row[0]: row[1] for row in by_source},
            "top_courts": {row[0]: row[1] for row in by_court},
            "top_gdpr_articles": {row[0]: row[1] for row in dsr_stats},
        }

        return Response(
            content=json.dumps(data, ensure_ascii=False, indent=2),
            media_type="application/json",
            headers={"Content-Disposition": "attachment; filename=dsr_statistics.json"},
        )

    else:  # CSV
        output = StringIO()
        output.write("\ufeff")  # BOM

        writer = csv.writer(output, delimiter=";")

        # Übersicht
        writer.writerow(["Statistik", "Wert"])
        writer.writerow(["Gesamt-Entscheidungen", total])
        writer.writerow([])

        # Nach Quelle
        writer.writerow(["Quelle", "Anzahl"])
        for row in by_source:
            writer.writerow(row)
        writer.writerow([])

        # Top Gerichte
        writer.writerow(["Gericht", "Anzahl"])
        for row in by_court:
            writer.writerow(row)
        writer.writerow([])

        # Top DSGVO-Artikel
        writer.writerow(["DSGVO-Artikel", "Anzahl"])
        for row in dsr_stats:
            writer.writerow(row)

        output.seek(0)

        return StreamingResponse(
            output,
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": "attachment; filename=dsr_statistics.csv"},
        )
